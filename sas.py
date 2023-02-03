import shutil
import openpyxl
import os
from pypsexec.client import Client
from winrmcp import Client

# # source_path = r'\\itech51\share51'
# source_path = r'\\operator4\c$\zDistr'
# dest_path = r'C:\Users\Dorofeev.E.BOOKCENTRE\Desktop\ssPyQt5_report_creation'
# file_name = '\\sas.txt'

# shutil.copyfile(source_path + file_name, dest_path + file_name)
# print(source_path + file_name, dest_path + file_name)

# создание файла excel
# boba = 'boba'
# filepath = rf"C:\Users\Dorofeev.E.BOOKCENTRE\Desktop\ssPyQt5_report_creation\{boba}.xlsx"
# wb = openpyxl.Workbook()
# wb.save(filepath)

# # # удаление файла
# if os.path.exists(r'C:\Users\Dorofeev.E.BOOKCENTRE\Desktop\ssPyQt5_report_creation\boba.xlsx'):
#     os.remove(
#         r'C:\Users\Dorofeev.E.BOOKCENTRE\Desktop\ssPyQt5_report_creation\boba.xlsx')


# wb = openpyxl.load_workbook(f'addresses.xlsx')
# sheet = wb.active
# print(sheet.max_row)

# for i in range(1, sheet.max_row + 1):
#     print(sheet[f'A{i}'].value)
#     # print(i)


shop_pc = '172.16.39.68'
dst_path = rf'\\\\{shop_pc}\c$'
login = '.\администратор'
password = 'qwerty-bc'


# # mount_command = rf'net use /user:{login} "{dst_path}" {password}'
# # mount_result = os.system(mount_command)
# # mount_command = rf'net localgroup Администраторы'
# # mount_result = os.system(mount_command)

# c = Client(shop_pc, username=login, password=password)
# print('connect ', c.connect())
# c.create_service()
# print(c.run_executable('notepad ', "notepad.exe"))
# print('cmd ', c.run_executable("cmd.exe", arguments="ipconfig /all"))
# c.disconnect()

os.system(rf'psexec \\172.16.11.212 -u Администратор -p qwerty-bc ipconfig')
# os.system(rf'net localgroup Пользователи')
